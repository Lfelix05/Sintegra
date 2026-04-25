import pandas as pd
import requests
import time
from openpyxl import load_workbook

def consultar_cnpj(cnpj):
    cnpj = "".join(filter(str.isdigit, str(cnpj)))
    if len(cnpj) != 14:
        return "CNPJ Inválido"
    url = f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            dados = response.json()
            return "Simples Nacional" if dados.get("opcao_pelo_simples") else "Normal"
        elif response.status_code == 404:
            return "CNPJ não encontrado"
        else:
            return "Erro na consulta"
    except requests.exceptions.Timeout:
        return "Timeout"
    except Exception:
        return "Erro de Conexão"

df = pd.read_excel("clientes.xlsx", dtype={"CNPJ": str})
total = len(df)
resultados = []

for i, cnpj in enumerate(df['CNPJ'], start=1):
    regime = consultar_cnpj(cnpj)
    resultados.append(regime)
    print(f"[{i}/{total}] CNPJ {cnpj} -> {regime}")
    time.sleep(0.2)

df['Regime Tributário'] = resultados
df["CNPJ"] = df["CNPJ"].apply(lambda x: "".join(filter(str.isdigit, str(x))).zfill(14))
df.to_excel("clientes.xlsx", index=False, engine="openpyxl")

wb = load_workbook("clientes.xlsx")
ws = wb.active
col_idx = list(df.columns).index("CNPJ") + 1
for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
    for cell in row:
        cell.number_format = "@"
wb.save("clientes.xlsx")
print("\nPlanilha atualizada com sucesso!")